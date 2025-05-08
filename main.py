import sys, re
import io
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import json

atribute_json = 'atribute.json'
#Print tiếng việt
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
with open(atribute_json, 'r', encoding='utf-8') as f:
            mappings = json.load(f)
#last_element = list(mappings["Trường data"])
#print(last_element[2]) # Print the last element of the JSON file
# Print both the key and value

"""
def fill_page_hang_hoa(filepath, formated, attribute_json_path):
    try:
        # Load attribute mappings
        with open(attribute_json_path, 'r', encoding='utf-8') as f:
            mappings = json.load(f)
        
        # Load the workbooks
        file_khach_hang = load_workbook(filepath)
        file_format = load_workbook(formated)

        # Find the correct sheet in source file
        source_sheet = None
        for sheet_name in file_khach_hang.sheetnames:
            for variation in mappings["Trường data"]["Hàng hóa"]["Hàng hóa"]:
                if sheet_name.lower() == variation.lower():
                    source_sheet = file_khach_hang[sheet_name]
                    break
            if source_sheet:
                break
        
        if not source_sheet:
            raise ValueError("No matching 'Hàng hóa' sheet found in source file")

        
        # Get target sheet
        trang_hang_hoa_format = file_format['Hàng hóa']
        
        # Find columns in source sheet using attribute mappings
        header_row = next(source_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        
        hang_hoa_fields = mappings["Trường data"]["Hàng hóa"].keys()
        # Convert to list and remove the first element which is the sheet name itself
        field_names = list(hang_hoa_fields)[1:]
        
        field_columns = {}  # This will store our field:index mappings
    
        for idx, header in enumerate(header_row, start=1):
            if header:
                header_lower = str(header).lower()
                # Check for field variations
                for field in mappings["Trường data"]["Hàng hóa"]:
                    for variation in mappings["Trường data"]["Hàng hóa"][field]:
                        if header_lower == variation.lower():
                            field_columns[field] = idx
                            break  # Break after first match is found
        
        if (len(field_columns) < len(field_names)):
            missing = [field for field in field_names if field not in field_columns]
            # Check for missing fields
            raise ValueError(f"Required columns not found: {', '.join(missing)}")

        # Find columns in target sheet
        target_header_row = next(trang_hang_hoa_format.iter_rows(min_row=1, max_row=1, values_only=True))
        
        field_columns_targeted = {}
        
        for idx, header in enumerate(target_header_row, start=1):
            if header:
                header_str = str(header).strip()
                if header_str in field_names:
                    field_columns_targeted[header_str] = idx
                
        # Copy data from source to target
        for row_idx in range(2, source_sheet.max_row + 1):
            row_data = [source_sheet.cell(row=row_idx, column=field_columns[field]).value for field in field_names]
            # Write to target sheet
            for field, col_idx in field_columns_targeted.items():
                trang_hang_hoa_format.cell(row=row_idx, column=col_idx).value = row_data[field_names.index(field)]


        # Save the workbook
        file_format.save(formated)
        print(f"Successfully copied data from {filepath} to {formated}")
        
    except Exception as e:
        print(f"Error: {str(e)}")
        raise
def fill_page_tai_xe(filepath, formated, attribute_json_path):
    try:
        # Load attribute mappings
        with open(attribute_json_path, 'r', encoding='utf-8') as f:
            mappings = json.load(f)
        
        # Load the workbooks
        file_khach_hang = load_workbook(filepath)
        file_format = load_workbook(formated)

        # Find the correct sheet in source file
        source_sheet = None
        for sheet_name in file_khach_hang.sheetnames:
            for variation in mappings["Trường data"]["Tài xế"]["Tài xế"]:
                if sheet_name.lower() == variation.lower():
                    source_sheet = file_khach_hang[sheet_name]
                    break
            if source_sheet:
                break
        
        if not source_sheet:
            raise ValueError("No matching 'Tài xế' sheet found in source file")

        
        # Get target sheet
        trang_tai_xe_format = file_format['Tài xế']
        
        # Find columns in source sheet using attribute mappings
        header_row = next(source_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        
        tai_xe_fields = mappings["Trường data"]["Tài xế"].keys()
        # Convert to list and remove the first element which is the sheet name itself
        field_names = list(tai_xe_fields)[1:]
        
        field_columns = {}  # This will store our field:index mappings
    
        for idx, header in enumerate(header_row, start=1):
            if header:
                header_lower = str(header).lower()
                # Check for field variations
                for field in mappings["Trường data"]["Tài xế"]:
                    for variation in mappings["Trường data"]["Tài xế"][field]:
                        if header_lower == variation.lower():
                            field_columns[field] = idx
                            break  # Break after first match is found
        
        if (len(field_columns) < len(field_names)):
            missing = [field for field in field_names if field not in field_columns]
            # Check for missing fields
            raise ValueError(f"Required columns not found: {', '.join(missing)}")

        # Find columns in target sheet
        target_header_row = next(trang_tai_xe_format.iter_rows(min_row=1, max_row=1, values_only=True))
        
        field_columns_targeted = {}
        
        for idx, header in enumerate(target_header_row, start=1):
            if header:
                header_str = str(header).strip()
                if header_str in field_names:
                    field_columns_targeted[header_str] = idx
                
        # Copy data from source to target
        for row_idx in range(2, source_sheet.max_row + 1):
            row_data = [source_sheet.cell(row=row_idx, column=field_columns[field]).value for field in field_names]
            # Write to target sheet
            for field, col_idx in field_columns_targeted.items():
                trang_tai_xe_format.cell(row=row_idx, column=col_idx).value = row_data[field_names.index(field)]


        # Save the workbook
        file_format.save(formated)
        print(f"Successfully copied data from {filepath} to {formated}")
        

    except Exception as e:
        print(f"Error: {str(e)}")
        raise
"""

def fill_pages(filepath, formated, attribute_json_path):
    try:
        # Load attribute mappings
        with open(attribute_json_path, 'r', encoding='utf-8') as f:
            mappings = json.load(f)
        formated_sheet_names = list(mappings["Trường data"])
        # Load the workbooks
        file_khach_hang = load_workbook(filepath)
        file_format = load_workbook(formated)

        for sheet_name in formated_sheet_names:
            process_sheet(file_khach_hang, file_format, mappings, sheet_name)

        # Process Hàng hóa if present
        #process_sheet(file_khach_hang, file_format, mappings, "Hàng hóa")
        
        # Process Tài xế if present
        #process_sheet(file_khach_hang, file_format, mappings, "Tài xế")

        # Save the workbook
        file_format.save(formated)
        print(f"Successfully copied data from {filepath} to {formated}")
        
    except Exception as e:
        print(f"Error: {str(e)}")
        raise

def process_sheet(source_wb, target_wb, mappings, sheet_type):
    header_row = None
    # Find the correct sheet in source file
    source_sheet = None
    for sheet_name in source_wb.sheetnames:
        if sheet_type in mappings["Trường data"]:
            for variation in mappings["Trường data"][sheet_type][sheet_type]:
                if sheet_name.lower() == variation.lower():
                    source_sheet = source_wb[sheet_name]
                    break
        if source_sheet:
            break
    
    if not source_sheet:
        print(f"No matching '{sheet_type}' sheet found in source file - skipping")
        return
    
    try:
        # Get target sheet
        target_sheet = target_wb[sheet_type]
        print(target_sheet)
        
        # Find columns in source sheet using attribute mappings
        header_row = next(source_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        print(source_sheet)
        
        fields = mappings["Trường data"][sheet_type].keys()
        # Convert to list and remove the first element which is the sheet name itself
        field_names = list(fields)[1:]
        
        field_columns = {}  # This will store our field:index mappings
    
        for idx, header in enumerate(header_row, start=1):
            if header:
                header_lower = str(header).lower()
                # Check for field variations
                for field in mappings["Trường data"][sheet_type]:
                    for variation in mappings["Trường data"][sheet_type][field]:
                        if header_lower == variation.lower():
                            field_columns[field] = idx
                            break  # Break after first match is found
        
        if (len(field_columns) < len(field_names)):
            missing = [field for field in field_names if field not in field_columns]
            # Check for missing fields
            raise ValueError(f"Required columns not found in {sheet_type}: {', '.join(missing)}")

        # Find columns in target sheet
        target_header_row = next(target_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        
        field_columns_targeted = {}
        
        for idx, header in enumerate(target_header_row, start=1):
            if header:
                header_str = str(header).strip()
                if header_str in field_names:
                    field_columns_targeted[header_str] = idx
                
        # Copy data from source to target
        for row_idx in range(2, source_sheet.max_row + 1):
            row_data = [source_sheet.cell(row=row_idx, column=field_columns[field]).value 
                      for field in field_names]
            # Write to target sheet
            for field, col_idx in field_columns_targeted.items():
                target_sheet.cell(row=row_idx, column=col_idx).value = row_data[field_names.index(field)]
                
        print(f"Successfully processed {sheet_type} sheet")
        
    except Exception as e:
        print(f"Error processing {sheet_type} sheet: {str(e)}")
        raise

#Call

filepath = 'Raw.xlsx'
formated = 'Form.xlsx'
atribute_json = 'atribute.json'
fill_pages(filepath, formated, atribute_json)



