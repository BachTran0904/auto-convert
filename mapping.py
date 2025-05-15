import sys
import io
from openpyxl import load_workbook
import json
import argparse

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
#global variables
field_columns = {}
target_columns = {}
filepath = "Raw.xlsx"
formated = "Form.xlsx"

def get_target_columns(target_sheet, mappings):
    target_columns = {}
    target_header_row = next(target_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    for idx, header in enumerate(target_header_row, start=1):
        if header:
            header_str = str(header).strip()
            # Check all possible fields in mappings
            for category in mappings["Trường data"]:
                if header_str in category:
                    target_columns[category] = idx
                    break
    return target_columns
    #print(f"Target columns mapping: {target_columns}")

def load_mappings(attribute_json_path):
    try:
        with open(attribute_json_path, 'r', encoding='utf-8') as f:
            mappings = json.load(f)
        return mappings
    except Exception as e:
        print(f"Error loading attribute JSON: {str(e)}")
        raise

def find_data_sheets(source_wb):
    """Find all sheets with data (not empty) in the workbook"""
    data_sheets = []
    for sheet in source_wb:
        if sheet.max_row > 0:  # At least have header row
            data_sheets.append(sheet)
    return data_sheets

def map_fields_to_columns(source_sheet, mappings):
    field_columns = {}
    header_row = next(source_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    
    for idx, header in enumerate(header_row, start=1):
        if header:
            header_lower = str(header).strip().lower()
            for category in mappings["Trường data"]:
                for field in mappings["Trường data"][category]:
                    if header_lower == field.strip().lower():
                        field_columns[category] = idx
                        break
    return field_columns
    #print(f"Field columns mapping: {field_columns}")

def copy_data_to_target(source_sheet, target_sheet, mappings):
    # Map source columns to fields
    field_columns = map_fields_to_columns(source_sheet, mappings)
    source_header_row = next(source_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    #print(f"Source header row: {source_header_row}")

    
    # Map target columns to fields
    target_columns = get_target_columns(target_sheet, mappings)
    target_header_row = next(target_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    #print(f"Target header row: {target_header_row}")
    
    print(f"Source columns: {field_columns}")
    print(f"Target columns: {target_columns}")

    # Copy data row by row
    for row_idx in range(2, source_sheet.max_row + 1):  # Skip header
        target_row = target_sheet.max_row + 1
        for field, source_col in field_columns.items():
            if (field in target_header_row):
                target_col = target_columns[field]
                source_col = field_columns[field]
                target_sheet.cell(row=target_row, column=target_col).value = source_sheet.cell(row=row_idx, column=source_col).value

def process_workbooks(source_wb, target_wb, mappings):
    try:
        # Get all data sheets from both workbooks
        source_sheets = find_data_sheets(source_wb)
        target_sheets = find_data_sheets(target_wb)
        
        # Process each corresponding sheet pair
        for source_sheet in source_sheets:
            for target_sheet in target_sheets:
                print(f"Processing sheet pair: {source_sheet.title} -> {target_sheet.title}")
                copy_data_to_target(source_sheet, target_sheet, mappings)
        
                        
    except Exception as e:
        print(f"Error processing data: {str(e)}")
        raise

def mapping(filepath, formated, attribute_json_path):
    try:
        mappings = load_mappings(attribute_json_path)
        file_khach_hang = load_workbook(filepath)
        file_format = load_workbook(formated)

        process_workbooks(file_khach_hang, file_format, mappings)
        
        # Save and return the formatted file path
        file_format.save(formated)
        print(f"Successfully copied data from {filepath} to {formated}")
        return formated

    except Exception as e:
        print(f"Error: {str(e)}")
        raise

#mapping(filepath, formated, "attribute.json")
"""
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Map data from source Excel to formatted Excel based on JSON attributes")
    parser.add_argument("filepath", help="Path to the source Excel file")
    parser.add_argument("formated", help="Path to the formatted Excel file")
    parser.add_argument("atribute_json", help="Path to the attribute JSON file")
    args = parser.parse_args()
    
    # Run the mapping and get the formatted file path
    result_file = mapping(args.filepath, args.formated, args.atribute_json)
    print(f"Formatted file saved at: {result_file}")
"""