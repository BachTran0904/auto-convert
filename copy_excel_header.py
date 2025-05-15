import sys
import io
from openpyxl import load_workbook
import json

file = 'Form.xlsx'
format = 'Form_chaun.xlsx'

# Load the source workbook
source_wb = load_workbook(file)
# Create a new workbook for the output
output_wb = load_workbook(format)

# Copy each sheet's first row
for sheet_name in source_wb.sheetnames:
    source_sheet = source_wb[sheet_name]
    # Get the first row values
    first_row_values = next(source_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    
    # Create or get the corresponding sheet in output workbook
    if sheet_name in output_wb.sheetnames:
        output_sheet = output_wb[sheet_name]
        output_sheet.delete_rows(1, output_sheet.max_row)  # Clear existing data
    else:
        output_sheet = output_wb.create_sheet(sheet_name)
    
    # Write the first row to the output sheet
    output_sheet.append(first_row_values)

# Save the output workbook
output_wb.save(format)